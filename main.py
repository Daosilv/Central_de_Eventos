import tkinter as tk
from tkinter import ttk, messagebox
import os
import sys
import tempfile
import atexit
from datetime import datetime
import openpyxl
import traceback
import time
import queue
import threading

try:
    from PIL import Image, ImageTk
except ImportError:
    Image = None
    ImageTk = None

try:
    import psutil
except ImportError:
    messagebox.showerror("Biblioteca Faltando", "A biblioteca 'psutil' é necessária. Instale com: pip install psutil")
    sys.exit(1)

try:
    import win32com.client as win32
except ImportError:
    messagebox.showerror("Biblioteca Faltando", "A biblioteca 'pywin32' é necessária para a exportação em PDF. Instale com: pip install pywin32")
    sys.exit(1)

FONTE_TITULO = ("Helvetica", 22, "bold")
FONTE_BOTAO_MENU = ("Helvetica", 14)
CAMINHO_BD = r'C:\BD_APP_Cemig\Primeiro_banco.db'

class App(tk.Tk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        self.model = None
        self.logo_photo = None

        try:
            if Image is None or ImageTk is None:
                raise ImportError("A biblioteca Pillow não está instalada.")
            
            logo_path = r'C:\BD_APP_Cemig\plemt_logo_ao_lado.PNG'
            logo_image = Image.open(logo_path)
            self.logo_photo = ImageTk.PhotoImage(logo_image)
        
        except Exception:
            pass

        style = ttk.Style()
        style.theme_use('clam')
        
        self.after_idle(self.setup_database_and_app)

        self.ajustes_window = None 
        self.title("Plemt - Plataforma de Estudos de Média Tensão")
        
        self.geometry("1280x720") 
        self.minsize(1000, 600)

        # Centraliza a janela principal
        self.center_window(self)

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        
        self.frames = {}
        menu_frame = MenuFrame(container, self, logo=self.logo_photo)
        self.frames[MenuFrame.__name__] = menu_frame
        menu_frame.grid(row=0, column=0, sticky="nsew")
        menu_frame.tkraise()

    def center_window(self, window):
        """Centraliza uma janela (Tk ou Toplevel) na tela."""
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')

    def setup_database_and_app(self):
        try:
            from model import DatabaseModel
            from memorial_frame import MemorialCalculoFrame
            
            self.model = DatabaseModel(CAMINHO_BD)
            colunas_memorial = MemorialCalculoFrame.get_db_columns()
            self.model.create_tables(colunas_memorial)
            self.model.populate_initial_data() # Popula com dados iniciais se necessário
        except Exception as e:
            error_details = traceback.format_exc()
            messagebox.showerror(
                "Erro Crítico na Inicialização",
                f"Ocorreu um erro ao iniciar o aplicativo. Verifique as permissões da pasta "
                f"e o arquivo do banco de dados.\n\n"
                f"Erro: {e}\n\nDetalhes:\n{error_details}"
            )
            self.destroy()

    # --- MÉTODOS PARA ABA AJUDA ---
    def get_all_ajuda(self):
        return self.model.get_all_ajuda()
    def get_ajuda_by_id(self, ajuda_id):
        return self.model.get_ajuda_by_id(ajuda_id)
    def add_ajuda(self, titulo):
        return self.model.add_ajuda(titulo)
    def update_ajuda(self, ajuda_id, titulo, texto, anexos):
        return self.model.update_ajuda(ajuda_id, titulo, texto, anexos)
    def delete_ajuda(self, ajuda_id):
        return self.model.delete_ajuda(ajuda_id)

    # --- MÉTODOS PAR TEMPO MORTO ---
    def get_all_tempo_morto(self):
        return self.model.get_all_tempo_morto()
    def add_tempo_morto(self, data):
        return self.model.add_tempo_morto(data)
    def update_tempo_morto(self, config_id, data):
        return self.model.update_tempo_morto(config_id, data)
    def delete_tempo_morto(self, config_id):
        return self.model.delete_tempo_morto(config_id)

    # --- MÉTODOS PARA GRUPO 4 ---
    def get_all_grupo4(self):
        return self.model.get_all_grupo4()
    def add_grupo4(self, data):
        return self.model.add_grupo4(data)
    def update_grupo4(self, config_id, data):
        return self.model.update_grupo4(config_id, data)
    def delete_grupo4(self, config_id):
        return self.model.delete_grupo4(config_id)

    def abrir_modulo_memorial(self):
        from memorial_frame import MemorialCalculoFrame
        memorial_window = tk.Toplevel(self)
        memorial_window.title("Memorial de Cálculo")
        memorial_window.geometry("1536x864")
        memorial_frame = MemorialCalculoFrame(parent=memorial_window, controller=self)
        memorial_frame.pack(side="top", fill="both", expand=True)
        self.center_window(memorial_window)

    def abrir_modulo_visualizacao(self):
        from visualizacao_frame import VisualizacaoFrame
        visualizacao_window = tk.Toplevel(self)
        visualizacao_window.title("Visualização de Dados")
        visualizacao_window.geometry("1536x864")
        visualizacao_frame = VisualizacaoFrame(parent=visualizacao_window, controller=self)
        visualizacao_frame.pack(side="top", fill="both", expand=True)
        self.center_window(visualizacao_window)

    def abrir_modulo_ajustes(self):
        from ajustes_frame import AjustesFrame
        if self.ajustes_window is not None and self.ajustes_window.winfo_exists():
            messagebox.showinfo("Janela já aberta", "A janela de Ajustes já está em uso.", parent=self)
            self.ajustes_window.lift()
            return
        ajustes_window = tk.Toplevel(self)
        ajustes_window.title("Ajustes Pendentes")
        ajustes_window.geometry("1536x864")
        ajustes_window.resizable(True, True)
        ajustes_frame = AjustesFrame(parent=ajustes_window, controller=self)
        ajustes_frame.pack(side="top", fill="both", expand=True)
        self.ajustes_window = ajustes_window
        self.ajustes_window.protocol("WM_DELETE_WINDOW", self.on_ajustes_close)
        self.center_window(ajustes_window)

    def on_ajustes_close(self):
        if self.ajustes_window:
            self.ajustes_window.destroy()
        self.ajustes_window = None

    def abrir_modulo_gestao(self):
        try:
            from gestao_frame import GestaoFrame
            gestao_window = tk.Toplevel(self)
            gestao_window.title("Gestão de Opções")
            gestao_window.geometry("950x600")
            # ALTERAÇÃO 5: Linhas  que a janela não bloqueie as outras
            # gestao_window.transient(self)
            # gestao_window.grab_set()
            gestao_frame = GestaoFrame(parent=gestao_window, controller=self)
            gestao_frame.pack(side="top", fill="both", expand=True)
            self.center_window(gestao_window)
        except ImportError:
            messagebox.showerror("Erro", "O módulo de Gestão (gestao_frame.py) não foi encontrado.")
        
    def salvar_rascunho(self, view, form_data):
        self.model.save_draft(form_data)
        messagebox.showinfo("Sucesso", "Rascunho salvo/atualizado com sucesso!", parent=view.parent_window)
        self.atualizar_listas_memorial(view)

    def salvar_e_enviar(self, view, form_data):
        self.model.save_final_and_manage_lists(form_data)
        messagebox.showinfo("Sucesso", "Ficha salva e enviada com sucesso!", parent=view.parent_window)
        view.limpar_formulario()
        self.atualizar_listas_memorial(view)
        
    def atualizar_listas_memorial(self, view):
        view.lista_ajustes_completa = self.model.get_recent_equipments_with_status()
        view.lista_rascunhos_completa = self.model.get_all_draft_equipments()
        view.reconstruir_lista_ajustes()
        view.reconstruir_lista_rascunhos()

    def export_to_excel(self, view, map_id):
        
        def show_wait_window(app_controller, parent, message):
            wait_window = tk.Toplevel(parent)
            wait_window.title("Aguarde")
            wait_window.transient(parent)
            wait_window.grab_set()
            wait_window.resizable(False, False)
            tk.Label(wait_window, text=message).pack(padx=30, pady=20)
            app_controller.center_window(wait_window)
            wait_window.update()
            return wait_window

        def worker(q, template_path, save_path, mapeamento, data):
            try:
                workbook = openpyxl.load_workbook(template_path, keep_vba=True)
                
                campo_condicao = 'condição'
                if campo_condicao in mapeamento:
                    aba, celula = mapeamento.pop(campo_condicao)
                    if aba in workbook.sheetnames:
                        sheet = workbook[aba]
                        valor = data.get(campo_condicao, "")
                        sheet[celula] = valor
                
                campos_cabecalho_processados = []
                for campo, (aba, celula) in list(mapeamento.items()):
                    if not campo.startswith(('g1_', 'g2_', 'g3_')):
                        if aba in workbook.sheetnames:
                            sheet = workbook[aba]
                            valor = data.get(campo, "")
                            sheet[celula] = valor
                        campos_cabecalho_processados.append(campo)
                
                for campo in campos_cabecalho_processados:
                    mapeamento.pop(campo)
                    
                for campo, (aba, celula) in mapeamento.items():
                    if aba in workbook.sheetnames:
                        sheet = workbook[aba]
                        valor = data.get(campo, "")
                        sheet[celula] = valor

                workbook.save(save_path)
                q.put({'status': 'success', 'path': save_path})

            except FileNotFoundError:
                q.put({'status': 'error', 'type': 'FileNotFound', 'path': template_path})
            except Exception as e:
                error_details = traceback.format_exc()
                q.put({'status': 'error', 'type': 'Exception', 'error': e, 'details': error_details})

        template_dir = r'C:\Central_de_Eventos\Diretório Excel Padrão'
        save_dir = r'C:\Central_de_Eventos\Diretório Excel Final'
        template_filename = 'Planilha_MonoTrifasico_vs35_REDE.xlsm'
        template_path = os.path.join(template_dir, template_filename)

        try:
            os.makedirs(save_dir, exist_ok=True)
        except OSError as e:
            messagebox.showerror("Erro de Diretório", f"Não foi possível criar o diretório final:\n{e}", parent=view.parent_window)
            return

        data = view._get_form_data()
        eqpto = data.get("eqpto")
        if not eqpto:
            messagebox.showwarning("Atenção", "Preencha ao menos o campo 'Eqpto' para exportar.", parent=view.parent_window)
            return
            
        mapeamentos_db = self.model.get_mappings_for_map(map_id)
        if not mapeamentos_db:
            messagebox.showerror("Erro", "Nenhum mapeamento para o Excel foi encontrado no mapa selecionado.", parent=view.parent_window)
            return
        
        mapeamento = {item['campo_app']: (item['aba_excel'], item['celula_excel']) for item in mapeamentos_db}

        data_str = datetime.now().strftime('%d%m%Y')
        new_filename = f"{eqpto}_{data_str}_MC.xlsm"
        save_path = os.path.join(save_dir, new_filename)

        loading_win = show_wait_window(self, view.parent_window, "Gerando planilha...")
        
        q = queue.Queue()
        thread = threading.Thread(target=worker, args=(q, template_path, save_path, mapeamento, data))
        thread.start()

        def check_queue():
            try:
                result = q.get_nowait()
                loading_win.destroy()
                if result['status'] == 'success':
                    messagebox.showinfo("Sucesso", f"Planilha salva em:\n{result['path']}", parent=view.parent_window)
                else:
                    if result['type'] == 'FileNotFound':
                        messagebox.showerror("Erro", f"O template da planilha não foi encontrado em:\n{result['path']}", parent=view.parent_window)
                    else:
                        messagebox.showerror("Erro Inesperado", f"Ocorreu um erro ao gerar a planilha:\n{result['error']}\n\nDetalhes:\n{result['details']}", parent=view.parent_window)
            except queue.Empty:
                view.parent_window.after(100, check_queue)

        view.parent_window.after(100, check_queue)
        
    def carregar_ficha_pelo_nome(self, view, eqpto_or_id, tabela, by_id=False):
        record_id = None
        if by_id:
            record_id = eqpto_or_id
        else:
            id_result = self.model.get_latest_id_for_equipment(eqpto_or_id, tabela)
            if id_result:
                record_id = id_result['id']

        if record_id:
            record_data = self.model.get_record_by_id(record_id, tabela)
            if record_data:
                view.preencher_formulario(record_data)

    def deletar_rascunhos(self, view, eqpto_list):
        self.model.delete_drafts(eqpto_list)
        messagebox.showinfo("Sucesso", f"{len(eqpto_list)} rascunhos foram deletados.", parent=view.parent_window)
        self.atualizar_listas_memorial(view)

    def carregar_ajustes_pendentes(self):
        return self.model.get_pending_adjustments()

    def concluir_ajuste(self, eqpto, status):
        self.model.complete_adjustment(eqpto, status)

    def carregar_observacoes(self, eqpto):
        return self.model.get_observations(eqpto)

    def adicionar_observacao(self, eqpto, autor, texto):
        self.model.add_observation(eqpto, autor, texto)

    def buscar_registros(self, filtros):
        return self.model.search_records(filtros)
    
    def get_latest_id_for_equipment(self, eqpto, table_name):
        return self.model.get_latest_id_for_equipment(eqpto, table_name)

    def buscar_todos_registros_visualizacao(self, filtros):
        return self.model.buscar_todos_registros_para_visualizacao(filtros)

    def get_record_by_id(self, record_id, table_name='cadastros'):
        return self.model.get_record_by_id(record_id, table_name)
        
    def update_ajuste_status(self, eqpto_list, status):
        self.model.update_adjustment_status(eqpto_list, status)

    def get_historico_ajustes(self):
        return self.model.get_adjustments_history()

    def get_categorias_gerais(self):
        return self.model.get_categorias_gerais()

    def get_opcoes_por_categoria_geral(self, categoria):
        return self.model.get_opcoes_por_categoria_geral(categoria)

    def adicionar_opcao_geral(self, categoria, valor):
        self.model.adicionar_opcao_geral(categoria, valor)

    def editar_opcao_geral(self, id_opcao, novo_valor):
        self.model.editar_opcao_geral(id_opcao, novo_valor)

    def deletar_opcao_geral(self, id_opcao):
        self.model.deletar_opcao_geral(id_opcao)
        
    def deletar_opcoes_por_categoria_geral(self, categoria):
        self.model.deletar_opcoes_por_categoria_geral(categoria)

    def get_opcoes_hierarquicas(self, id_pai=None):
        return self.model.get_opcoes_hierarquicas(id_pai)

    def adicionar_opcao_hierarquica(self, categoria, valor, id_pai=None):
        self.model.adicionar_opcao_hierarquica(categoria, valor, id_pai)

    def editar_opcao_hierarquica(self, id_opcao, novo_valor):
        self.model.editar_opcao_hierarquica(id_opcao, novo_valor)

    def deletar_opcao_hierarquica(self, id_opcao):
        self.model.deletar_opcao_hierarquica(id_opcao)

    def get_todos_cabos(self):
        return self.model.get_todos_cabos()
    
    def get_propriedades_cabo(self, trecho):
        return self.model.get_propriedades_cabo(trecho)

    def adicionar_cabo(self, trecho, nominal, admissivel):
        self.model.adicionar_cabo(trecho, nominal, admissivel)

    def editar_cabo(self, id_cabo, trecho, nominal, admissivel):
        self.model.editar_cabo(id_cabo, trecho, nominal, admissivel)

    def deletar_cabo(self, id_cabo):
        self.model.deletar_cabo(id_cabo)

    def get_all_map_names(self):
        return self.model.get_all_map_names()
        
    def add_map_name(self, name):
        self.model.add_map_name(name)
        
    def rename_map_name(self, map_id, new_name):
        self.model.rename_map_name(map_id, new_name)
        
    def delete_map(self, map_id):
        self.model.delete_map(map_id)
        
    def get_mappings_for_map(self, map_id):
        return self.model.get_mappings_for_map(map_id)
        
    def add_mappings_to_map(self, map_id, mappings_list):
        self.model.add_mappings_to_map(map_id, mappings_list)

    def update_mapping(self, mapping_id, data):
        self.model.update_mapping(mapping_id, data)

    def deletar_mapeamento(self, map_id):
        self.model.deletar_mapeamento(map_id)
        
    def deletar_mapeamentos_por_mapa(self, map_id):
        self.model.deletar_mapeamentos_por_mapa(map_id)
        
    def get_all_users(self):
        return self.model.get_all_users()

    def add_user(self, data):
        return self.model.add_user(data)

    def update_user(self, user_id, data):
        return self.model.update_user(user_id, data)

    def delete_user(self, user_id):
        return self.model.delete_user(user_id)


class MenuFrame(tk.Frame):
    def __init__(self, parent, controller, logo=None):
        super().__init__(parent)
        self.controller = controller
        
        # Cores do novo design
        cor_fundo_principal = "white"  # Fundo branco
        cor_fundo_botao = "#a7c957"      # Verde claro
        cor_borda_botao = "#386641"      # Verde escuro
        cor_texto_botao = "#ffffff"      # Branco
        cor_botao_hover = "#c2e08a"      # Verde mais claro para hover

        self.configure(bg=cor_fundo_principal)

        # --- Top Frame (para o logo e título) ---
        # Ajustar para que o logo fique no topo e ocupe menos espaço vertical
        top_frame = tk.Frame(self, bg=cor_fundo_principal)
        top_frame.pack(side="top", fill="x", pady=10) # Pack instead of place for better flow

        title_frame = tk.Frame(top_frame, bg=cor_fundo_principal)
        title_frame.pack(expand=True) # Center the title/logo horizontally

        if logo:
            logo_label = tk.Label(title_frame, image=logo, bg=cor_fundo_principal)
            logo_label.pack() # No pady here, let top_frame handle it
        else:
            fallback_text = "Plemt - Plataforma de Estudos de Média Tensão"
            lbl_title = tk.Label(title_frame, text=fallback_text, font=FONTE_TITULO, bg=cor_fundo_principal, fg="#a7c957", pady=20)
            lbl_title.pack()

        # --- Main content frame for buttons ---
        # Este frame ocupará o restante do espaço e centralizará os botões
        content_frame = tk.Frame(self, bg=cor_fundo_principal)
        content_frame.pack(side="top", fill="both", expand=True)

        buttons_frame = tk.Frame(content_frame, bg=cor_fundo_principal)
        buttons_frame.pack(expand=True) # Center buttons vertically and horizontally
        
        style = ttk.Style()
        style.configure("Menu.TButton", 
                        font=FONTE_BOTAO_MENU, 
                        padding=20, 
                        background=cor_fundo_botao, 
                        foreground=cor_texto_botao,
                        borderwidth=2,
                        bordercolor=cor_borda_botao,
                        relief="raised")
        
        style.map("Menu.TButton",
                  background=[('active', cor_botao_hover)],
                  relief=[('pressed', 'sunken')])
        
        ttk.Button(buttons_frame, text="Visualização", style="Menu.TButton", command=self.controller.abrir_modulo_visualizacao).grid(row=0, column=0, padx=20, pady=20)
        ttk.Button(buttons_frame, text="Memorial de Cálculo", style="Menu.TButton", command=self.controller.abrir_modulo_memorial).grid(row=0, column=1, padx=20, pady=20)
        ttk.Button(buttons_frame, text="Ajustes", style="Menu.TButton", command=self.controller.abrir_modulo_ajustes).grid(row=0, column=2, padx=20, pady=20)
        ttk.Button(buttons_frame, text="Gestão", style="Menu.TButton", command=self.controller.abrir_modulo_gestao).grid(row=0, column=3, padx=20, pady=20)

if __name__ == "__main__":
    lock_file_path = os.path.join(tempfile.gettempdir(), 'plemt.lock')
    current_script_path = os.path.abspath(sys.argv[0])

    def is_process_running(pid):
        if pid is None: return False
        try:
            process = psutil.Process(pid)
            if any(current_script_path in arg for arg in process.cmdline()):
                return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            return False
        return False

    def remover_lock_file():
        try:
            if os.path.exists(lock_file_path): os.remove(lock_file_path)
        except OSError: pass

    if os.path.exists(lock_file_path):
        old_pid = None
        try:
            with open(lock_file_path, 'r') as f:
                content = f.read().strip()
                if content: old_pid = int(content)
        except (IOError, ValueError): old_pid = None
        if is_process_running(old_pid):
            messagebox.showerror("Aplicação já em execução", "A Plemt já está aberta.")
            sys.exit(1)
        else:
            remover_lock_file()

    try:
        with open(lock_file_path, 'w') as f: f.write(str(os.getpid()))
        atexit.register(remover_lock_file)
    except IOError:
        messagebox.showerror("Erro de Permissão", "Não foi possível criar o arquivo de lock.")
        sys.exit(1)

    app = App()
    app.mainloop()
